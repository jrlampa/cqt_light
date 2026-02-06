import openpyxl

file_path = r'C:\myworld\utils\PLANILHA_DESTRAVADA.xlsm'

def extract_formulas(path):
    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    
    sheets_to_check = ['Coeficiente Unitário', 'ANÁLISE PONTO A PONTO', 'QDT Dutra 2.3 Lado Esquerdo']
    
    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found.")
            continue
            
        print(f"\n--- Checking Formulas in {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Look for cells with formulas
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"Cell {cell.coordinate}: {cell.value}")
                    count += 1
                    if count >= 30: # Limit output
                        break
            if count >= 30:
                break

if __name__ == "__main__":
    extract_formulas(file_path)
