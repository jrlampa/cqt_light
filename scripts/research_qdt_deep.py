import openpyxl
import json

file_path = r'C:\\myworld\\utils\\PLANILHA_DESTRAVADA.xlsm'

def extract_qdt_deep_logic(path):
    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    ws = wb["QDT Dutra 2.3 Lado Esquerdo"]
    
    # Headers research (Rows 10-12 usually contain headers)
    headers = {}
    for col in range(1, 110):
        col_letter = openpyxl.utils.get_column_letter(col)
        val = ws.cell(row=10, column=col).value
        if val:
            headers[col_letter] = str(val)

    # Core logic extraction for a specific row (e.g. row 15)
    row_logic = {}
    for col in range(1, 110):
        col_letter = openpyxl.utils.get_column_letter(col)
        cell = ws.cell(row=15, column=col)
        if cell.value:
            row_logic[col_letter] = {
                "val": str(cell.value),
                "header": headers.get(col_letter, "")
            }

    # Final result components
    final_components = {
        "C33": ws["C33"].value,
        "CX102": ws["CX102"].value,
        "CW104": ws["CW104"].value,
        "CW100": ws["CW100"].value,
        "CV105": ws["CV105"].value,
        "BV4": ws["BV4"].value
    }

    results = {
        "headers_found": headers,
        "row_15_logic": row_logic,
        "final_components": final_components
    }

    with open('qdt_deep_logic.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4, ensure_ascii=False)
    
    print("Deep research complete.")

if __name__ == "__main__":
    extract_qdt_deep_logic(file_path)
