import openpyxl
import pandas as pd
import json

file_path = r'C:\\myworld\\utils\\PLANILHA_DESTRAVADA.xlsm'

def extract_workflow_logic(path):
    print(f"Loading workbook: {path}")
    # data_only=False to get formulas
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    
    results = {
        "distrib_cargas": {},
        "qdt_esquerdo": {},
        "ramais": {}
    }
    
    # 1. Check "Distrib. Cargas"
    if "Distrib. Cargas" in wb.sheetnames:
        ws = wb["Distrib. Cargas"]
        results["distrib_cargas"]["G3"] = ws["G3"].value
        # Look for cells that depend on G3 or define the distribution
        for row in range(1, 40):
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str) and "G3" in cell.value:
                    results["distrib_cargas"][f"{cell.coordinate}"] = cell.value
    
    # 2. Check "QDT Dutra 2.3 Lado Esquerdo" (assuming this is the one)
    target_qdt = "QDT Dutra 2.3 Lado Esquerdo"
    if target_qdt in wb.sheetnames:
        ws = wb[target_qdt]
        results["qdt_esquerdo"]["C33"] = ws["C33"].value
        results["qdt_esquerdo"]["CX102"] = ws["CX102"].value
        
        # Look for how C33 is calculated
        # The user mentioned C33 reads from CX102
        
        # Sample some rows to see the pattern of the circuit path
        for row in range(10, 30):
            row_data = {
                "Trecho": ws.cell(row=row, column=1).value, # A
                "kVA": ws.cell(row=row, column=3).value,   # C
                "Seção": ws.cell(row=row, column=5).value, # E
                "Formula_QD": ws.cell(row=row, column=102).value # CX is 102? wait, CX is a column index.
            }
            # Column CX is column 102 in Excel? A=1, Z=26, AA=27...
            # CX: C=3, X=24 -> 3*26 + 24 = 78 + 24 = 102. Correct.
            results["qdt_esquerdo"][f"row_{row}"] = row_data

    # 3. Check "ramais"
    if "ramais" in wb.sheetnames:
        df_ramais = pd.read_excel(path, sheet_name="ramais", nrows=10)
        results["ramais"]["columns"] = df_ramais.columns.tolist()
        results["ramais"]["sample"] = df_ramais.iloc[0:3].to_dict()

    with open('workflow_research.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4, ensure_ascii=False)
    
    print("Research complete. results saved to workflow_research.json")

if __name__ == "__main__":
    extract_workflow_logic(file_path)
