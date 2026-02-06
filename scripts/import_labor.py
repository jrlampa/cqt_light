"""
CQT Light V2 - Labor Cost Importer
Imports labor costs from XLSM files in PLANILHA CUSTO MODULAR folder.
Each file represents a different partner company and regional.

Structure: Sheet 'Modular Aéreo Urbano' contains the cost data
- Column A: Service description
- Column B: Service code (SAP-like)
- Column C: Mestre de Serviço cost
- Column D: Valor Unitário Serv. (main cost)
"""

import sqlite3
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    import openpyxl

# Paths
BASE_DIR = Path(__file__).parent.parent
XLSM_DIR = BASE_DIR / "PLANILHA CUSTO MODULAR"
DB_PATH = BASE_DIR / "frontend" / "cqt_light.db"


def extract_partner_info(filename):
    """Extract partner name and regional from filename."""
    # Example: CM AÉREO URBANO - 01.07.2025 - CONTRATO 4600009167 - EMPRESA ELLCA - REGIONAL VALE.xlsm
    partner = "UNKNOWN"
    regional = "UNKNOWN"
    contract = "UNKNOWN"
    
    # Extract EMPRESA
    match = re.search(r'EMPRESA\s+(\w+)', filename)
    if match:
        partner = match.group(1)
    
    # Extract REGIONAL
    match = re.search(r'REGIONAL\s+([\w\s]+)\.xlsm', filename, re.IGNORECASE)
    if match:
        regional = match.group(1).strip()
    
    # Extract CONTRATO
    match = re.search(r'CONTRATO\s+(\d+)', filename)
    if match:
        contract = match.group(1)
    
    return partner, regional, contract


def import_xlsm_file(filepath, conn):
    """Import labor costs from a single XLSM file."""
    filename = filepath.name
    partner, regional, contract = extract_partner_info(filename)
    
    print(f"\n📂 Processing: {filename}")
    print(f"   Partner: {partner}, Regional: {regional}, Contract: {contract}")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=True, data_only=True)
    except Exception as e:
        print(f"   ⚠️ Error loading file: {e}")
        return 0
    
    # Find the cost sheet
    cost_sheet = None
    for sheet_name in wb.sheetnames:
        if 'modular' in sheet_name.lower() or 'aéreo' in sheet_name.lower():
            cost_sheet = wb[sheet_name]
            break
    
    if not cost_sheet:
        print(f"   ⚠️ No cost sheet found. Available: {wb.sheetnames}")
        return 0
    
    cursor = conn.cursor()
    count = 0
    
    # Iterate through rows looking for data
    for row_idx, row in enumerate(cost_sheet.iter_rows(min_row=5, values_only=True), start=5):
        if not row or len(row) < 4:
            continue
        
        descricao = row[0]
        codigo = row[1]
        
        # Skip header rows and invalid data
        if not descricao or not codigo:
            continue
        
        # Skip if description looks like a header
        if any(x in str(descricao).upper() for x in ['INSTALAÇÃO DE REDE', 'CUSTO MODULAR', 'NÚMERO DA NOTA']):
            continue
        
        # Try to parse the code and price
        try:
            # Code should be numeric or a valid code format
            codigo_str = str(codigo).strip()
            if not codigo_str or codigo_str.lower() in ['none', 'nan', 'mestre']:
                continue
            
            # Get price from column C or D (whichever has valid data)
            preco = None
            for col in [2, 3]:  # Columns C and D (0-indexed: 2, 3)
                if col < len(row) and row[col]:
                    try:
                        preco = float(row[col])
                        break
                    except (ValueError, TypeError):
                        continue
            
            if preco is None or preco <= 0:
                continue
            
            # Create unique code including partner info
            codigo_mo = f"{codigo_str}_{partner}_{regional}".replace(" ", "_")
            descricao_str = str(descricao).strip()
            
            # Insert or update
            cursor.execute("""
                INSERT INTO mao_de_obra (codigo_mo, descricao, unidade, preco_bruto)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(codigo_mo) DO UPDATE SET
                    descricao = excluded.descricao,
                    preco_bruto = excluded.preco_bruto
            """, (codigo_mo, descricao_str, 'UN', preco))
            count += 1
            
        except Exception as e:
            # Skip invalid rows silently
            pass
    
    conn.commit()
    print(f"   ✅ Imported {count} labor cost entries")
    return count


def main():
    print("=" * 60)
    print("CQT Light V2 - Labor Cost Importer")
    print("=" * 60)
    print(f"\nSource: {XLSM_DIR}")
    print(f"Database: {DB_PATH}\n")
    
    if not XLSM_DIR.exists():
        print(f"❌ Directory not found: {XLSM_DIR}")
        return
    
    # Find all XLSM files (exclude temp files starting with ~$)
    xlsm_files = [f for f in XLSM_DIR.glob("*.xlsm") if not f.name.startswith("~$")]
    print(f"Found {len(xlsm_files)} XLSM files to process\n")
    
    if not xlsm_files:
        print("❌ No XLSM files found")
        return
    
    # Connect to database
    conn = sqlite3.connect(DB_PATH)
    
    total_count = 0
    for filepath in xlsm_files:
        count = import_xlsm_file(filepath, conn)
        total_count += count
    
    # Show final stats
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM mao_de_obra")
    mo_count = cursor.fetchone()[0]
    
    print("\n" + "=" * 60)
    print(f"📊 Import Summary:")
    print(f"   Files processed: {len(xlsm_files)}")
    print(f"   Total entries imported: {total_count}")
    print(f"   Total mão de obra in DB: {mo_count}")
    print("=" * 60)
    
    conn.close()


if __name__ == "__main__":
    main()
