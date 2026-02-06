"""
CQT Light V3 - Complete Data Seeder
Seeds the new simplified schema with materials, kits, and servicos_cm.
"""

import json
import sqlite3
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    import openpyxl

BASE_DIR = Path(__file__).parent.parent
DB_PATH = BASE_DIR / "frontend" / "cqt_light.db"
DATA_DIR = BASE_DIR / "data"
XLSM_DIR = BASE_DIR / "PLANILHA CUSTO MODULAR"

def create_schema(conn):
    """Create fresh tables."""
    cursor = conn.cursor()
    
    # Drop old tables
    cursor.execute("DROP TABLE IF EXISTS kit_composicao")
    cursor.execute("DROP TABLE IF EXISTS kit_servicos")
    cursor.execute("DROP TABLE IF EXISTS mao_de_obra")
    cursor.execute("DROP TABLE IF EXISTS kits")
    cursor.execute("DROP TABLE IF EXISTS materiais")
    cursor.execute("DROP TABLE IF EXISTS servicos_cm")
    conn.commit()
    
    # Create new schema
    schema = """
    CREATE TABLE materiais (
      sap TEXT PRIMARY KEY,
      descricao TEXT NOT NULL,
      unidade TEXT DEFAULT 'UN',
      preco_unitario REAL DEFAULT 0
    );

    CREATE TABLE servicos_cm (
      codigo TEXT PRIMARY KEY,
      descricao TEXT NOT NULL,
      preco_bruto REAL DEFAULT 0
    );

    CREATE TABLE kits (
      codigo_kit TEXT PRIMARY KEY,
      descricao_kit TEXT NOT NULL,
      codigo_servico TEXT,
      custo_servico REAL DEFAULT 0
    );

    CREATE TABLE kit_composicao (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      codigo_kit TEXT NOT NULL,
      sap TEXT NOT NULL,
      quantidade REAL DEFAULT 1,
      UNIQUE(codigo_kit, sap)
    );

    CREATE INDEX idx_kit_composicao_kit ON kit_composicao(codigo_kit);
    CREATE INDEX idx_kit_composicao_sap ON kit_composicao(sap);
    """
    cursor.executescript(schema)
    conn.commit()
    print("✅ Schema created")


def import_materials(conn):
    """Import materials from JSON catalog."""
    catalog_path = DATA_DIR / "catalog" / "material_catalog.json"
    if not catalog_path.exists():
        print(f"⚠️ Material catalog not found: {catalog_path}")
        return 0
    
    with open(catalog_path, 'r', encoding='utf-8') as f:
        materials = json.load(f)
    
    cursor = conn.cursor()
    count = 0
    for sap, data in materials.items():
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO materiais (sap, descricao, unidade, preco_unitario)
                VALUES (?, ?, ?, ?)
            """, (
                sap,
                data.get('description', ''),
                data.get('unit', 'UN'),
                float(data.get('price', 0) or 0)
            ))
            count += 1
        except Exception as e:
            pass
    
    conn.commit()
    print(f"✅ Imported {count} materials")
    return count


def import_kits(conn):
    """Import kits and compositions from JSON."""
    kits_path = DATA_DIR / "kits" / "kits.json"
    if not kits_path.exists():
        print(f"⚠️ Kits file not found: {kits_path}")
        return 0
    
    with open(kits_path, 'r', encoding='utf-8') as f:
        kits = json.load(f)
    
    cursor = conn.cursor()
    kit_count = 0
    comp_count = 0
    
    for codigo, data in kits.items():
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO kits (codigo_kit, descricao_kit, codigo_servico, custo_servico)
                VALUES (?, ?, NULL, 0)
            """, (codigo, data.get('name', codigo)))
            kit_count += 1
            
            for mat in data.get('materials', []):
                sap = str(mat.get('sap', '')).strip()
                qty = float(mat.get('qty', 1) or 1)
                if sap:
                    cursor.execute("""
                        INSERT OR REPLACE INTO kit_composicao (codigo_kit, sap, quantidade)
                        VALUES (?, ?, ?)
                    """, (codigo, sap, qty))
                    comp_count += 1
        except Exception as e:
            pass
    
    conn.commit()
    print(f"✅ Imported {kit_count} kits with {comp_count} compositions")
    return kit_count


def import_servicos(conn):
    """Import services from XLSM files."""
    if not XLSM_DIR.exists():
        print(f"⚠️ XLSM directory not found: {XLSM_DIR}")
        return 0
    
    xlsm_files = [f for f in XLSM_DIR.glob("*.xlsm") if not f.name.startswith("~$")]
    if not xlsm_files:
        print("⚠️ No XLSM files found")
        return 0
    
    cursor = conn.cursor()
    count = 0
    
    for filepath in xlsm_files:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=True, data_only=True)
            
            # Find cost sheet
            sheet = None
            for name in wb.sheetnames:
                if 'modular' in name.lower() or 'aéreo' in name.lower():
                    sheet = wb[name]
                    break
            
            if not sheet:
                continue
            
            for row in sheet.iter_rows(min_row=5, max_row=500, values_only=True):
                if not row or len(row) < 5:
                    continue
                
                descricao = row[0]
                codigo = row[1]
                
                if not descricao or not codigo:
                    continue
                
                codigo_str = str(codigo).strip()
                if not codigo_str.isdigit():
                    continue
                
                # Get price from column C, D, or E
                preco = None
                for col in [2, 3, 4]:
                    if col < len(row) and row[col]:
                        try:
                            preco = float(row[col])
                            break
                        except:
                            continue
                
                if preco is None or preco <= 0:
                    continue
                
                cursor.execute("""
                    INSERT OR REPLACE INTO servicos_cm (codigo, descricao, preco_bruto)
                    VALUES (?, ?, ?)
                """, (codigo_str, str(descricao).strip()[:200], preco))
                count += 1
                
        except Exception as e:
            print(f"⚠️ Error processing {filepath.name}: {e}")
    
    conn.commit()
    print(f"✅ Imported {count} service entries")
    return count


def main():
    print("=" * 60)
    print("CQT Light V3 - Data Seeder")
    print("=" * 60)
    print(f"\nDatabase: {DB_PATH}\n")
    
    # Remove old database
    if DB_PATH.exists():
        DB_PATH.unlink()
        print("🗑️ Removed old database")
    
    conn = sqlite3.connect(DB_PATH)
    
    create_schema(conn)
    import_materials(conn)
    import_kits(conn)
    import_servicos(conn)
    
    # Stats
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM materiais")
    mat_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM kits")
    kit_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM servicos_cm")
    srv_count = cursor.fetchone()[0]
    
    print("\n" + "=" * 60)
    print("📊 Final Stats:")
    print(f"   Materiais: {mat_count:,}")
    print(f"   Kits: {kit_count:,}")
    print(f"   Serviços: {srv_count:,}")
    print("=" * 60)
    
    conn.close()


if __name__ == "__main__":
    main()
